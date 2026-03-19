from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def build_workbook(output_path: Path) -> None:
    wb = Workbook()
    ws_txn = wb.active
    ws_txn.title = "Transactions"

    headers = ["Date", "Business", "Type", "Category", "Amount (INR)", "Month"]
    ws_txn.append(headers)

    data = [
        ("2026-03-18", "Cafe", "Income", "Sales", 14250),
        ("2026-03-18", "Travel", "Expense", "Fuel", 3200),
        ("2026-03-17", "Textile", "Income", "Wholesale", 22100),
        ("2026-03-17", "Hardware", "Expense", "Inventory", 8900),
        ("2026-03-16", "Cafe", "Expense", "Supplies", 2800),
        ("2026-03-15", "Travel", "Income", "Packages", 18100),
        ("2026-03-14", "Hardware", "Income", "Retail", 26400),
        ("2026-03-14", "Textile", "Expense", "Logistics", 11900),
        ("2026-02-28", "Cafe", "Income", "Sales", 12600),
        ("2026-02-25", "Travel", "Expense", "Marketing", 4600),
        ("2026-02-19", "Textile", "Income", "Wholesale", 19700),
        ("2026-02-13", "Hardware", "Expense", "Rent", 7100),
        ("2026-01-26", "Cafe", "Expense", "Utilities", 2300),
        ("2026-01-20", "Travel", "Income", "Packages", 15300),
        ("2026-01-15", "Textile", "Expense", "Payroll", 8600),
        ("2026-01-09", "Hardware", "Income", "Retail", 21000),
        ("2025-12-26", "Cafe", "Income", "Sales", 11800),
        ("2025-12-17", "Travel", "Expense", "Maintenance", 5100),
        ("2025-12-10", "Textile", "Income", "Bulk", 17350),
        ("2025-12-02", "Hardware", "Expense", "Inventory", 6200),
    ]

    for i, row in enumerate(data, start=2):
        ws_txn.cell(i, 1, row[0])
        ws_txn.cell(i, 2, row[1])
        ws_txn.cell(i, 3, row[2])
        ws_txn.cell(i, 4, row[3])
        ws_txn.cell(i, 5, row[4])
        ws_txn.cell(i, 6, f'=TEXT(A{i},"mmm-yy")')

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws_txn[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in ws_txn.iter_rows(min_row=2, max_row=1 + len(data), min_col=1, max_col=1):
        for cell in row:
            cell.number_format = "dd-mmm-yyyy"
    for row in ws_txn.iter_rows(min_row=2, max_row=1 + len(data), min_col=5, max_col=5):
        for cell in row:
            cell.number_format = '[Red](₹#,##0);₹#,##0;"-"'

    widths = [14, 14, 12, 15, 16, 12]
    for idx, width in enumerate(widths, start=1):
        ws_txn.column_dimensions[get_column_letter(idx)].width = width

    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "Expense & Profit Tracker - Advanced Summary"
    ws_sum["A1"].font = Font(size=14, bold=True, color="1F4E78")
    ws_sum.merge_cells("A1:H1")

    kpi_fill = PatternFill("solid", fgColor="E2F0D9")
    section_fill = PatternFill("solid", fgColor="D9E1F2")
    label_font = Font(bold=True, color="1F4E78")

    ws_sum["A3"] = "KPI"
    ws_sum["B3"] = "Value"
    for ref in ("A3", "B3"):
        ws_sum[ref].font = header_font
        ws_sum[ref].fill = header_fill

    ws_sum["A4"] = "Total Revenue"
    ws_sum["A5"] = "Total Expenses"
    ws_sum["A6"] = "Net Profit"
    ws_sum["A7"] = "Profit Margin %"
    ws_sum["A8"] = "Monthly Growth % (Mar vs Feb 2026)"

    ws_sum["B4"] = '=SUMIFS(Transactions!E:E,Transactions!C:C,"Income")'
    ws_sum["B5"] = '=SUMIFS(Transactions!E:E,Transactions!C:C,"Expense")'
    ws_sum["B6"] = "=B4-B5"
    ws_sum["B7"] = "=IFERROR(B6/B4,0)"
    ws_sum["B8"] = (
        '=IFERROR((SUMIFS(Transactions!E:E,Transactions!F:F,"Mar-26",Transactions!C:C,"Income")-'
        'SUMIFS(Transactions!E:E,Transactions!F:F,"Mar-26",Transactions!C:C,"Expense")-'
        '(SUMIFS(Transactions!E:E,Transactions!F:F,"Feb-26",Transactions!C:C,"Income")-'
        'SUMIFS(Transactions!E:E,Transactions!F:F,"Feb-26",Transactions!C:C,"Expense")))/'
        'ABS(SUMIFS(Transactions!E:E,Transactions!F:F,"Feb-26",Transactions!C:C,"Income")-'
        'SUMIFS(Transactions!E:E,Transactions!F:F,"Feb-26",Transactions!C:C,"Expense")),0)'
    )

    for row in range(4, 9):
        ws_sum[f"A{row}"].font = label_font
        ws_sum[f"B{row}"].fill = kpi_fill

    for ref in ("B4", "B5", "B6"):
        ws_sum[ref].number_format = '[Red](₹#,##0);₹#,##0;"-"'
    for ref in ("B7", "B8"):
        ws_sum[ref].number_format = "0.00%"

    ws_sum["D3"] = "Business Comparison"
    ws_sum["D3"].font = header_font
    ws_sum["D3"].fill = header_fill
    ws_sum.merge_cells("D3:G3")

    ws_sum["D4"] = "Business"
    ws_sum["E4"] = "Income (INR)"
    ws_sum["F4"] = "Expense (INR)"
    ws_sum["G4"] = "Profit (INR)"
    for ref in ("D4", "E4", "F4", "G4"):
        ws_sum[ref].font = header_font
        ws_sum[ref].fill = section_fill

    businesses = ["Cafe", "Travel", "Textile", "Hardware"]
    for i, business in enumerate(businesses, start=5):
        ws_sum[f"D{i}"] = business
        ws_sum[f"E{i}"] = f'=SUMIFS(Transactions!E:E,Transactions!B:B,D{i},Transactions!C:C,"Income")'
        ws_sum[f"F{i}"] = f'=SUMIFS(Transactions!E:E,Transactions!B:B,D{i},Transactions!C:C,"Expense")'
        ws_sum[f"G{i}"] = f"=E{i}-F{i}"
        for col in ("E", "F", "G"):
            ws_sum[f"{col}{i}"].number_format = '[Red](₹#,##0);₹#,##0;"-"'

    ws_sum["D11"] = "Monthly Trend"
    ws_sum["D11"].font = header_font
    ws_sum["D11"].fill = header_fill
    ws_sum.merge_cells("D11:G11")

    ws_sum["D12"] = "Month"
    ws_sum["E12"] = "Income (INR)"
    ws_sum["F12"] = "Expense (INR)"
    ws_sum["G12"] = "Profit (INR)"
    for ref in ("D12", "E12", "F12", "G12"):
        ws_sum[ref].font = header_font
        ws_sum[ref].fill = section_fill

    months = ["Dec-25", "Jan-26", "Feb-26", "Mar-26"]
    for i, month in enumerate(months, start=13):
        ws_sum[f"D{i}"] = month
        ws_sum[f"E{i}"] = f'=SUMIFS(Transactions!E:E,Transactions!F:F,D{i},Transactions!C:C,"Income")'
        ws_sum[f"F{i}"] = f'=SUMIFS(Transactions!E:E,Transactions!F:F,D{i},Transactions!C:C,"Expense")'
        ws_sum[f"G{i}"] = f"=E{i}-F{i}"
        for col in ("E", "F", "G"):
            ws_sum[f"{col}{i}"].number_format = '[Red](₹#,##0);₹#,##0;"-"'

    ws_sum["A11"] = "Expense by Category"
    ws_sum["A11"].font = header_font
    ws_sum["A11"].fill = header_fill
    ws_sum.merge_cells("A11:B11")
    ws_sum["A12"] = "Category"
    ws_sum["B12"] = "Expense (INR)"
    for ref in ("A12", "B12"):
        ws_sum[ref].font = header_font
        ws_sum[ref].fill = section_fill

    categories = ["Fuel", "Inventory", "Supplies", "Logistics", "Marketing", "Rent", "Utilities", "Payroll", "Maintenance"]
    for i, category in enumerate(categories, start=13):
        ws_sum[f"A{i}"] = category
        ws_sum[f"B{i}"] = f'=SUMIFS(Transactions!E:E,Transactions!D:D,A{i},Transactions!C:C,"Expense")'
        ws_sum[f"B{i}"].number_format = '[Red](₹#,##0);₹#,##0;"-"'

    for col, width in {"A": 34, "B": 26, "D": 14, "E": 16, "F": 16, "G": 16}.items():
        ws_sum.column_dimensions[col].width = width

    thin = Side(style="thin", color="D1D5DB")
    for row in ws_sum.iter_rows(min_row=3, max_row=20, min_col=1, max_col=7):
        for cell in row:
            if cell.value is not None:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                if cell.column in (2, 5, 6, 7):
                    cell.alignment = Alignment(horizontal="right")

    ws_ch = wb.create_sheet("Charts")
    ws_ch["A1"] = "Visual Dashboard"
    ws_ch["A1"].font = Font(size=14, bold=True, color="1F4E78")

    line = LineChart()
    line.title = "Profit Trend Over Time"
    line.y_axis.title = "INR"
    line.x_axis.title = "Month"
    line_data = Reference(ws_sum, min_col=7, min_row=12, max_row=16)
    line_cats = Reference(ws_sum, min_col=4, min_row=13, max_row=16)
    line.add_data(line_data, titles_from_data=True)
    line.set_categories(line_cats)
    line.height = 7
    line.width = 12
    ws_ch.add_chart(line, "A3")

    pie = PieChart()
    pie.title = "Expense Category Mix"
    pie_labels = Reference(ws_sum, min_col=1, min_row=13, max_row=21)
    pie_data = Reference(ws_sum, min_col=2, min_row=12, max_row=21)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_labels)
    pie.height = 7
    pie.width = 10
    ws_ch.add_chart(pie, "M3")

    bar = BarChart()
    bar.title = "Business Income vs Expense"
    bar.y_axis.title = "INR"
    bar.x_axis.title = "Business"
    bar_data = Reference(ws_sum, min_col=5, max_col=6, min_row=4, max_row=8)
    bar_cats = Reference(ws_sum, min_col=4, min_row=5, max_row=8)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_cats)
    bar.height = 8
    bar.width = 12
    ws_ch.add_chart(bar, "A20")

    wb.save(output_path)


if __name__ == "__main__":
    out_dir = Path(__file__).resolve().parents[1] / "output" / "spreadsheet"
    out_dir.mkdir(parents=True, exist_ok=True)
    output_path = out_dir / "Expense_Profit_Tracker_Advanced.xlsx"
    build_workbook(output_path)
    print(output_path)
